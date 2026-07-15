"""Excel (.xlsx) report writer.

Produces a multi-sheet workbook from the final AgentState:

  Summary           -> run metadata + disclaimer
  Picks             -> the ranked FinalPick list with thesis / risks / tax notes
  Factor breakdown  -> per-ticker factor scores grid (the quant heart)
  Snapshots         -> raw per-ticker numerical fields (the source of truth)
  Analyst signals   -> each AnalystSignal as a row
  Debate            -> bull/bear turns
  Notes             -> risk + tax notes

Implementation note: this is openpyxl-only (no pandas dependency at runtime),
so the file is small and the code path is easy to reason about.
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import reports_dir
from ..graph.state import AgentState


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)


def _style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[row].height = 22


def _autosize(ws, min_w: int = 10, max_w: int = 60) -> None:
    """Cheap auto-size: set width to ~max content length per column."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_w
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            for line in str(v).splitlines() or [""]:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[letter].width = min(max_len + 2, max_w)


def _slug(text: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_-]+", "-", text.strip())
    return text.strip("-")[:60] or "report"


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def _write_summary(wb: Workbook, state: AgentState) -> None:
    ws = wb.active
    ws.title = "Summary"

    profile = state.profile or {}
    rows = [
        ("Profile id", state.profile_id),
        ("Profile display", profile.get("display_name", state.profile_id)),
        ("Country / currency", f"{profile.get('country', '')} / {profile.get('currency', '')}"),
        ("Target", state.target),
        ("Universe", state.universe_name or profile.get("universe", {}).get("default", "")),
        ("Top N", state.top_n),
        ("Candidates considered", len(state.candidate_tickers)),
        ("Used LLM", state.use_llm),
        ("Debate rounds", state.max_debate_rounds),
        ("Generated at", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    ws.cell(row=1, column=1, value="Field").font = _HEADER_FONT
    ws.cell(row=1, column=2, value="Value").font = _HEADER_FONT
    _style_header(ws, 1, 2)
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=str(k)).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v if isinstance(v, (int, float, str, bool)) else str(v))

    # Factor weights table.
    weights = profile.get("factor_weights", {}) or {}
    if weights:
        start = len(rows) + 4
        ws.cell(row=start, column=1, value="Factor").font = _HEADER_FONT
        ws.cell(row=start, column=2, value="Weight").font = _HEADER_FONT
        _style_header(ws, start, 2)
        for j, (factor, w) in enumerate(weights.items(), start=start + 1):
            ws.cell(row=j, column=1, value=factor)
            ws.cell(row=j, column=2, value=float(w)).number_format = "0.00"

    # Disclaimer at the bottom.
    last = ws.max_row + 2
    ws.cell(
        row=last,
        column=1,
        value=(
            "Disclaimer: educational/research output. Not financial advice. No "
            "real orders are placed. Data is from public free sources and may "
            "be stale or incomplete. Verify before acting."
        ),
    ).alignment = _WRAP_ALIGN
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=2)
    ws.row_dimensions[last].height = 60

    _autosize(ws)


def _write_picks(wb: Workbook, state: AgentState) -> None:
    ws = wb.create_sheet("Picks")
    headers = [
        "Rank",
        "Ticker",
        "Name",
        "Composite",
        "Profile fit",
        "Coverage",
        "Factor std-dev",
        "Confidence",
        "FX exposure",
        "Expected gross return",
        "Expected after-tax return",
        "Suggested horizon",
        "Thesis",
        "Key risks",
        "Tax / profile notes",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for p in state.picks:
        ws.append(
            [
                p.rank,
                p.ticker,
                p.name or "",
                p.composite_score if p.composite_score is not None else None,
                p.profile_fit if p.profile_fit is not None else None,
                p.coverage if p.coverage is not None else None,
                p.factor_std_dev if p.factor_std_dev is not None else None,
                p.confidence,
                "yes" if p.is_cross_currency else "no",
                p.expected_gross_return,
                p.expected_after_tax_return,
                p.suggested_horizon or "",
                p.thesis or "",
                "\n".join(p.key_risks or []),
                "\n".join(p.tax_notes or []),
            ]
        )

    # Number formatting
    for row_idx in range(2, ws.max_row + 1):
        for col in (4, 5, 7, 8):
            ws.cell(row=row_idx, column=col).number_format = "0.00"
        ws.cell(row=row_idx, column=6).number_format = "0.0%"
        for col in (10, 11):
            ws.cell(row=row_idx, column=col).number_format = "0.0%"
        for col in (13, 14, 15):
            ws.cell(row=row_idx, column=col).alignment = _WRAP_ALIGN

    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_factor_breakdown(wb: Workbook, state: AgentState) -> None:
    ws = wb.create_sheet("Factor breakdown")
    factor_names = ["quality", "value", "momentum", "financial_health", "earnings_quality"]
    headers = [
        "Rank",
        "Ticker",
        "Name",
        "Sector",
        "Composite",
        "Raw composite",
        *factor_names,
        "Coverage",
        "Coverage weight",
        "Profile fit",
        "Factor std-dev",
        "Floor breaches",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, rep in enumerate(state.factor_reports, start=1):
        scores = rep.get("factor_scores", {}) or {}
        ws.append(
            [
                i,
                rep.get("ticker"),
                rep.get("name") or "",
                rep.get("sector") or "",
                rep.get("composite_score"),
                rep.get("raw_composite"),
                *(scores.get(f) for f in factor_names),
                rep.get("coverage"),
                rep.get("coverage_weight"),
                rep.get("profile_fit"),
                rep.get("factor_std_dev"),
                "; ".join(rep.get("floor_breaches") or []),
            ]
        )

    last_score_col = 6 + len(factor_names)  # composite + raw + factors
    for row_idx in range(2, ws.max_row + 1):
        for col in range(5, last_score_col + 1):
            ws.cell(row=row_idx, column=col).number_format = "0.00"
        ws.cell(row=row_idx, column=last_score_col + 1).number_format = "0.0%"  # coverage
        ws.cell(row=row_idx, column=last_score_col + 2).number_format = "0.00"  # coverage weight
        ws.cell(row=row_idx, column=last_score_col + 3).number_format = "0.00"  # profile fit
        ws.cell(row=row_idx, column=last_score_col + 4).number_format = "0.00"  # std dev

    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_snapshots(wb: Workbook, state: AgentState) -> None:
    ws = wb.create_sheet("Snapshots")
    fields = [
        "ticker",
        "name",
        "currency",
        "sector",
        "industry",
        "country",
        "market_cap",
        "price",
        "pe_trailing",
        "pb",
        "ps",
        "ev_to_ebitda",
        "earnings_yield",
        "fcf_yield",
        "dividend_yield",
        "roe",
        "roic",
        "gross_margin",
        "operating_margin",
        "profit_margin",
        "debt_to_equity",
        "net_debt_to_ebitda",
        "current_ratio",
        "cash_conversion",
        "revenue_growth",
        "earnings_growth",
        "momentum_12_1",
        "momentum_6_1",
        "volatility_annualized",
        "beta",
        "data_agreement",
        "data_sources",
        "fetch_status",
        "is_cross_currency",
    ]
    ws.append(fields)
    _style_header(ws, 1, len(fields))

    for snap in state.snapshots:
        row = []
        for f in fields:
            v = snap.get(f)
            if isinstance(v, list):
                v = ", ".join(str(x) for x in v)
            row.append(v)
        ws.append(row)

    # Format numeric columns; market_cap as 0, others 0.0000
    if ws.max_row > 1:
        for col_idx, fname in enumerate(fields, start=1):
            if fname in {"market_cap"}:
                fmt = "#,##0"
            elif fname in {"price"}:
                fmt = "0.00"
            elif fname in {
                "ticker",
                "name",
                "currency",
                "sector",
                "industry",
                "country",
            }:
                fmt = None
            else:
                fmt = "0.0000"
            if fmt:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt

    ws.freeze_panes = "B2"
    _autosize(ws, max_w=24)


def _write_analyst_signals(wb: Workbook, state: AgentState) -> None:
    ws = wb.create_sheet("Analyst signals")
    headers = ["Role", "Ticker", "Stance", "Score", "Confidence", "Rationale", "Evidence"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for s in state.analyst_signals:
        ws.append(
            [
                s.role,
                s.ticker or "",
                s.stance,
                s.score,
                s.confidence,
                s.rationale or "",
                "\n".join(s.evidence or []),
            ]
        )

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=4).number_format = "0.00"
        ws.cell(row=row_idx, column=5).number_format = "0.00"
        ws.cell(row=row_idx, column=6).alignment = _WRAP_ALIGN
        ws.cell(row=row_idx, column=7).alignment = _WRAP_ALIGN

    ws.freeze_panes = "A2"
    _autosize(ws, max_w=80)


def _write_debate(wb: Workbook, state: AgentState) -> None:
    if not state.debate:
        return
    ws = wb.create_sheet("Debate")
    headers = ["Round", "Side", "Text"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for turn in state.debate:
        ws.append([turn.round_idx + 1, turn.side, turn.text or ""])
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).alignment = _WRAP_ALIGN
        ws.row_dimensions[row_idx].height = 80

    ws.freeze_panes = "A2"
    _autosize(ws, max_w=100)


def _write_data_health(wb: Workbook, state: AgentState) -> None:
    h = state.data_health or {}
    if not h:
        return
    ws = wb.create_sheet("Data health")
    ws.append(["Field", "Value"])
    _style_header(ws, 1, 2)
    rows = [
        ("Severity", h.get("severity")),
        ("Tickers requested", h.get("requested")),
        ("Tickers fetched", h.get("fetched")),
        ("Tickers failed", h.get("failed")),
        ("Average data coverage", h.get("avg_coverage")),
        ("Average cross-source agreement", h.get("avg_agreement")),
        ("Sources used", ", ".join(h.get("sources_used") or [])),
        ("Dropouts", ", ".join(h.get("dropouts") or [])),
    ]
    for k, v in rows:
        ws.append([k, v])
    if h.get("messages"):
        ws.append([])
        ws.append(["Messages", ""])
        for m in h["messages"]:
            ws.append(["", m])

    if h.get("low_agreement"):
        ws.append([])
        ws.append(["Ticker", "Agreement"])
        for t, a in h["low_agreement"]:
            ws.append([t, a])

    # Format the agreement / coverage cells.
    for row_idx in range(2, ws.max_row + 1):
        c = ws.cell(row=row_idx, column=2)
        if c.value is not None and isinstance(c.value, float) and -1 <= c.value <= 1:
            c.number_format = "0.0%"
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).alignment = _WRAP_ALIGN
    _autosize(ws, max_w=80)


def _write_factor_regime(wb: Workbook, state: AgentState) -> None:
    fr = state.factor_regime or {}
    if not fr:
        return
    ws = wb.create_sheet("Factor regime")
    ws.append(["Factor", "Top-Bottom 12-1m spread"])
    _style_header(ws, 1, 2)
    for k, v in (fr.get("factor_returns") or {}).items():
        ws.append([k, v])
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).number_format = "0.00%"
    if fr.get("regime_warnings"):
        ws.append([])
        ws.append(["Warning", ""])
        for w in fr["regime_warnings"]:
            ws.append([w, ""])
    _autosize(ws, max_w=80)


def _write_notes(wb: Workbook, state: AgentState) -> None:
    ws = wb.create_sheet("Notes")
    ws.append(["Type", "Note"])
    _style_header(ws, 1, 2)
    for n in state.risk_notes:
        ws.append(["risk", n])
    for n in state.tax_notes:
        ws.append(["tax", n])
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).alignment = _WRAP_ALIGN
    _autosize(ws, max_w=100)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_excel_report(state: AgentState) -> Path:
    """Write a multi-sheet xlsx workbook for the given state. Returns the path."""
    wb = Workbook()

    _write_summary(wb, state)
    _write_data_health(wb, state)
    _write_factor_regime(wb, state)
    _write_picks(wb, state)
    _write_factor_breakdown(wb, state)
    _write_snapshots(wb, state)
    _write_analyst_signals(wb, state)
    _write_debate(wb, state)
    _write_notes(wb, state)

    folder = reports_dir()
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    fname = f"{stamp}-{state.profile_id}-{_slug(state.target)}.xlsx"
    out_path = folder / fname
    wb.save(out_path)
    return out_path
