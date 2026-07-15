"""Light-weight walk-forward backtest of the deterministic factor composite.

Approach (intentionally simple):
- Pull daily prices for each ticker via DataProvider (yfinance).
- Quarterly rebalance: at the end of each quarter, rank tickers by a *price-
  only proxy* for the composite (12-1m momentum + recent 6-1m momentum +
  inverse volatility) since true point-in-time fundamentals require a paid
  data feed. This is *not* the full factor composite but it's directionally
  representative of momentum/quality tilts and gives a free baseline.
- Equal-weight top-N for the next quarter, subtract a configurable
  transaction cost on rebalance, and apply the profile's long-term capital
  gains tax rate to the realised gain at the end of each year.
- Track equity curve, monthly returns, Sharpe, Sortino, max drawdown.

The backtest writes an Excel workbook with: Summary, EquityCurve, Holdings.

Caveats logged on the Summary sheet:
- No fundamentals - this is a price-only proxy.
- No survivorship adjustment - if a ticker delisted in the window we just
  drop it.
- Fixed equal weight, no risk parity.
- Single starting universe (the profile's seed/live constituents) - no
  historical re-membership tracking.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterable

from ..data.provider import DataProvider


@dataclass
class BacktestConfig:
    profile_id: str
    profile: dict
    tickers: list[str]
    start: str = "2020-01-01"
    top_n: int = 10
    transaction_cost_bps: float = 15.0
    benchmark: str | None = None      # e.g. "^NSEI" or "^GDAXI" - optional


@dataclass
class BacktestResult:
    equity_curve: list[tuple[str, float]] = field(default_factory=list)
    holdings_log: list[tuple[str, list[str]]] = field(default_factory=list)
    benchmark_curve: list[tuple[str, float]] = field(default_factory=list)
    metrics: dict = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)

    def to_excel(self, path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "Backtest Summary"
        ws["A1"].font = Font(bold=True, size=14)
        row = 3
        ws.cell(row=row, column=1, value="Metric").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
        for k, v in self.metrics.items():
            row += 1
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
        row += 2
        ws.cell(row=row, column=1, value="Notes").font = Font(bold=True)
        for n in self.notes:
            row += 1
            ws.cell(row=row, column=1, value=n)

        eq = wb.create_sheet("EquityCurve")
        eq.append(["date", "portfolio", "benchmark"])
        bench_map = dict(self.benchmark_curve)
        for d, v in self.equity_curve:
            eq.append([d, v, bench_map.get(d)])

        hl = wb.create_sheet("Holdings")
        hl.append(["rebalance_date", "tickers"])
        for d, lst in self.holdings_log:
            hl.append([d, ", ".join(lst)])

        wb.save(path)


def _to_quarter_end_indices(dates: list[str]) -> list[int]:
    """Return indices in `dates` whose month transitions to a quarter end."""
    out = []
    last_q = None
    for i, d in enumerate(dates):
        try:
            dt_ = datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            continue
        q = (dt_.year, (dt_.month - 1) // 3)
        if last_q is None:
            last_q = q
        elif q != last_q:
            out.append(i - 1)  # the last day of the previous quarter
            last_q = q
    return out


def _annualised(returns: list[float], periods_per_year: int) -> float:
    if not returns:
        return 0.0
    mean = sum(returns) / len(returns)
    return mean * periods_per_year


def _sharpe(returns: list[float], periods_per_year: int) -> float:
    if len(returns) < 2:
        return 0.0
    mean = sum(returns) / len(returns)
    var = sum((r - mean) ** 2 for r in returns) / (len(returns) - 1)
    sd = var**0.5
    if sd == 0:
        return 0.0
    return (mean / sd) * (periods_per_year**0.5)


def _sortino(returns: list[float], periods_per_year: int) -> float:
    if len(returns) < 2:
        return 0.0
    mean = sum(returns) / len(returns)
    downs = [r for r in returns if r < 0]
    if not downs:
        return 0.0
    dd_var = sum(r * r for r in downs) / len(downs)
    dd_sd = dd_var**0.5
    if dd_sd == 0:
        return 0.0
    return (mean / dd_sd) * (periods_per_year**0.5)


def _max_drawdown(curve: list[float]) -> float:
    if not curve:
        return 0.0
    peak = curve[0]
    mdd = 0.0
    for v in curve:
        if v > peak:
            peak = v
        dd = (v - peak) / peak if peak else 0.0
        if dd < mdd:
            mdd = dd
    return mdd


def run_backtest(cfg: BacktestConfig) -> BacktestResult:
    provider = DataProvider(use_stooq=False)  # backtest is heavy enough
    res = BacktestResult()
    res.notes.append(
        "Price-only proxy for the composite (momentum + inv-vol). "
        "No point-in-time fundamentals used."
    )
    res.notes.append("No survivorship adjustment; delisted names are silently dropped.")
    res.notes.append(
        f"Round-trip cost: {cfg.transaction_cost_bps} bps applied at every rebalance."
    )
    tax_rate = float(((cfg.profile or {}).get("tax") or {}).get("long_term_rate", 0.0))
    res.notes.append(
        f"Long-term tax rate {tax_rate*100:.2f}% applied at year-end on realized gain."
    )

    # Pull ~5 years of history per ticker.
    histories: dict[str, list[dict]] = {}
    for t in cfg.tickers:
        rows = provider.get_history(t, period="5y")
        if rows:
            histories[t] = rows

    if not histories:
        res.notes.append("No price data available - aborting.")
        return res

    # Build a unified date axis (intersection of dates across tickers).
    date_sets = [set(r["date"] for r in rows) for rows in histories.values()]
    common_dates = sorted(set.intersection(*date_sets))
    if not common_dates:
        res.notes.append("No common dates across tickers - aborting.")
        return res

    # Filter by start date.
    common_dates = [d for d in common_dates if d >= cfg.start]
    if len(common_dates) < 60:
        res.notes.append("Too few common dates after start filter - aborting.")
        return res

    # Per-ticker close map keyed by date (only dates in common_dates).
    closes: dict[str, dict[str, float]] = {}
    for t, rows in histories.items():
        m = {r["date"]: float(r["close"]) for r in rows if r["close"] > 0}
        closes[t] = {d: m[d] for d in common_dates if d in m}

    # Quarter-end rebalance dates.
    qe_idx = _to_quarter_end_indices(common_dates)
    if not qe_idx:
        res.notes.append("Could not compute quarter ends - aborting.")
        return res

    # Iterate.
    equity = 1.0
    equity_curve_dates: list[str] = []
    equity_curve_values: list[float] = []

    cost = cfg.transaction_cost_bps / 10000.0
    period_returns: list[float] = []
    yearly_pnl: dict[int, float] = {}
    held: list[str] = []

    last_idx = 0
    for q_end in qe_idx:
        # Score each ticker on the lookback up to q_end using a price-only
        # proxy: 12-1m return + 6-1m return - 0.5 * trailing volatility.
        scores: list[tuple[str, float]] = []
        for t in cfg.tickers:
            cmap = closes.get(t, {})
            hist_dates = [d for d in common_dates[: q_end + 1] if d in cmap]
            if len(hist_dates) < 252 + 21:
                continue
            recent = cmap[hist_dates[-21]]
            past12 = cmap[hist_dates[-(252 + 21)]]
            past6 = cmap[hist_dates[-(126 + 21)]] if len(hist_dates) >= 126 + 21 else past12
            mom12 = (recent / past12) - 1.0 if past12 else 0.0
            mom6 = (recent / past6) - 1.0 if past6 else 0.0
            # Recent 60-day vol.
            window = hist_dates[-60:]
            vals = [cmap[d] for d in window]
            rets = [
                vals[i] / vals[i - 1] - 1.0
                for i in range(1, len(vals))
                if vals[i - 1] > 0
            ]
            if not rets:
                continue
            mean = sum(rets) / len(rets)
            var = sum((r - mean) ** 2 for r in rets) / max(len(rets) - 1, 1)
            sd = (var * 252) ** 0.5
            score = mom12 + 0.5 * mom6 - 0.3 * sd
            scores.append((t, score))

        scores.sort(key=lambda x: x[1], reverse=True)
        new_held = [t for t, _ in scores[: cfg.top_n]]

        # Apply transaction cost on rebalance (one-way for dropped, one-way for added).
        turnover = (
            len(set(new_held).symmetric_difference(set(held))) / max(2 * cfg.top_n, 1)
        )
        equity *= 1 - turnover * cost
        held = new_held
        res.holdings_log.append((common_dates[q_end], list(held)))

        # Track equity through the next quarter.
        next_q_end = qe_idx[qe_idx.index(q_end) + 1] if qe_idx.index(q_end) + 1 < len(qe_idx) else len(common_dates) - 1
        # Equal-weighted portfolio return between q_end -> next_q_end.
        if held:
            tickers_with_data = [
                t for t in held
                if common_dates[q_end] in closes[t] and common_dates[next_q_end] in closes[t]
            ]
            if tickers_with_data:
                rets = [
                    closes[t][common_dates[next_q_end]]
                    / closes[t][common_dates[q_end]]
                    - 1.0
                    for t in tickers_with_data
                ]
                period_ret = sum(rets) / len(rets)
                period_returns.append(period_ret)
                # Apply year-end tax: track per-year cumulative gain, tax at year change.
                year = int(common_dates[next_q_end][:4])
                yearly_pnl[year] = yearly_pnl.get(year, 0.0) + period_ret * equity
                equity *= 1 + period_ret

        equity_curve_dates.append(common_dates[next_q_end])
        equity_curve_values.append(equity)
        last_idx = next_q_end

    # Year-end tax (best-effort, applied on the curve).
    if tax_rate > 0:
        for year, gain in yearly_pnl.items():
            if gain <= 0:
                continue
            tax_amount = gain * tax_rate
            # Find the last equity point in that year and reduce.
            for i, d in enumerate(equity_curve_dates):
                if int(d[:4]) == year and (
                    i + 1 == len(equity_curve_dates)
                    or int(equity_curve_dates[i + 1][:4]) != year
                ):
                    factor = 1.0 - (tax_amount / equity_curve_values[i]) if equity_curve_values[i] > 0 else 1.0
                    for j in range(i, len(equity_curve_values)):
                        equity_curve_values[j] *= factor
                    break

    # Optional benchmark.
    if cfg.benchmark:
        bench_rows = provider.get_history(cfg.benchmark, period="5y") or []
        bm = {r["date"]: float(r["close"]) for r in bench_rows if r["close"] > 0}
        # Index aligned to equity_curve_dates (forward-filled).
        bench_curve: list[tuple[str, float]] = []
        anchor = None
        last_v = None
        for d in equity_curve_dates:
            v = bm.get(d) or last_v
            if v is None:
                continue
            last_v = v
            if anchor is None:
                anchor = v
            bench_curve.append((d, v / anchor if anchor else 1.0))
        res.benchmark_curve = bench_curve

    res.equity_curve = list(zip(equity_curve_dates, equity_curve_values))

    # Metrics (quarterly returns, annualised).
    res.metrics = {
        "tickers_evaluated": len(cfg.tickers),
        "rebalances": len(period_returns),
        "final_equity": round(equity, 4),
        "total_return": round(equity - 1.0, 4),
        "annualised_return": round(_annualised(period_returns, 4), 4),
        "sharpe": round(_sharpe(period_returns, 4), 4),
        "sortino": round(_sortino(period_returns, 4), 4),
        "max_drawdown": round(_max_drawdown(equity_curve_values), 4),
        "transaction_cost_bps": cfg.transaction_cost_bps,
        "tax_rate": tax_rate,
        "start": cfg.start,
        "end": equity_curve_dates[-1] if equity_curve_dates else "",
        "benchmark": cfg.benchmark or "",
    }
    return res
