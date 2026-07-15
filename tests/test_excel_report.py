"""Tests for the Excel (.xlsx) report writer."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.config import load_profile
from src.graph.orchestrator import run as run_graph
from src.graph.state import AgentState


def test_excel_report_written_india(fake_provider, force_stub_llm):
    profile = load_profile("india_adult")
    state = AgentState(
        profile_id="india_adult",
        profile=profile,
        target="best IT stocks in India",
        top_n=5,
        use_llm=False,
        write_excel=True,
        max_debate_rounds=0,
    )
    out = run_graph(state)
    assert out.excel_path is not None
    p = Path(out.excel_path)
    assert p.exists() and p.suffix == ".xlsx" and p.stat().st_size > 0

    wb = load_workbook(p, read_only=True, data_only=True)
    expected = {"Summary", "Picks", "Factor breakdown", "Snapshots", "Analyst signals", "Notes"}
    assert expected.issubset(set(wb.sheetnames))

    # Picks sheet has at least one data row + header.
    picks = wb["Picks"]
    rows = list(picks.iter_rows(values_only=True))
    assert rows[0][:3] == ("Rank", "Ticker", "Name")
    assert len(rows) >= 2
    # The first data row should reference an .NS ticker (India profile).
    assert any(isinstance(r[1], str) and r[1].endswith(".NS") for r in rows[1:])

    # Notes sheet contains India tax language.
    notes = wb["Notes"]
    note_text = " ".join(
        str(c) for r in notes.iter_rows(values_only=True) for c in r if c
    )
    assert "LTCG" in note_text or "STCG" in note_text


def test_no_excel_flag_skips_xlsx(fake_provider, force_stub_llm):
    profile = load_profile("germany_student")
    state = AgentState(
        profile_id="germany_student",
        profile=profile,
        target="best banks in Germany",
        top_n=3,
        use_llm=False,
        write_excel=False,
        max_debate_rounds=0,
    )
    out = run_graph(state)
    assert out.excel_path is None
    # Markdown still written.
    assert out.report_path is not None and Path(out.report_path).exists()


def test_excel_germany_has_abgeltung(fake_provider, force_stub_llm):
    profile = load_profile("germany_student")
    state = AgentState(
        profile_id="germany_student",
        profile=profile,
        target="best banks in Germany",
        top_n=3,
        use_llm=False,
        write_excel=True,
        max_debate_rounds=0,
    )
    out = run_graph(state)
    p = Path(out.excel_path)
    wb = load_workbook(p, read_only=True, data_only=True)
    notes_text = " ".join(
        str(c) for r in wb["Notes"].iter_rows(values_only=True) for c in r if c
    )
    assert "Abgeltungssteuer" in notes_text or "Sparerpauschbetrag" in notes_text
